import torch
import numpy as np
import pandas as pd
import os

from generate import gen_multi, quant_multi

from model.RandomForestQR import QuantileForest
from model.LinearQR import LinearQR
from model.KernelQR import KernelQR
from model.SplineQR import SplineQR
from model.DQR import DQR
from model.NQNet import NQNet
from model.NC_QR_DQN import NC_QR_DQN
from model.DQRP import DQRP
from model.ChebNet import ChebNet

#%% Set the model and hyperparameters


model='scenario1';error='scenario1';d=6;SIZE=2**10; df=2
#model='scenario1';error='scenario1';d=6;SIZE=2**11; df=2
#model='scenario1';error='scenario1';d=6;SIZE=2**12; df=2


#model='scenario2';error='t';d=15;SIZE=2**10; df=3
#model='scenario2';error='t';d=15;SIZE=2**11; df=3
#model='scenario2';error='t';d=15;SIZE=2**12; df=3

#model='scenario3';error='sinex_t';d=30;SIZE=2**10; df=3
#model='scenario3';error='sinex_t';d=30;SIZE=2**11; df=3
#model='scenario3';error='sinex_t';d=30;SIZE=2**12; df=3

# You can uncomment other settings as needed

torch.manual_seed(2024)
A = -1 * torch.randint(0, 3, [d, 1]) * torch.randn([d, 1])
torch.manual_seed(2025)
B = -1 * torch.randint(0, 3, [d, 1]) * torch.randn([d, 1])

taus = torch.linspace(0.05, 0.95, 19).unsqueeze(1)

R = 5
width = 256
lr = 0.01

# List of methods to evaluate
method_names = ['DQR', 'NC', 'NQ', 'DQRP', 'cheb', 'QR', 'RF', 'KQR', 'SPL']
L1_dict = {name: torch.zeros([R, len(taus)]) for name in method_names}
L2_dict = {name: torch.zeros([R, len(taus)]) for name in method_names}

#%% Run the simulation

for r in range(R):
    print(f"Replication {r+1} Start.")
    data_train = gen_multi(A, B, size=SIZE, d=d, model=model, error=error, df=df)
    x_train, y_train = data_train[:][0], data_train[:][1]

    # Fit all models
    net_NQ = NQNet(quantiles=taus)
    net_NQ.fit(X=x_train.numpy(), y=y_train.numpy(), width_vec=[d, width, width, width, len(taus)], activation='ELU', lr=lr, epochs=100)
    net_NC = NC_QR_DQN(quantiles=taus)
    net_NC.fit(X=x_train.numpy(), y=y_train.numpy(), logit_layer=[d, width, width, width, len(taus)], factor_layer=[d, width, width, width, 2], activation='ReLU', epochs=100, lr=lr)
    net_DQR = DQR(quantiles=taus)
    net_DQR.fit(x_train.numpy(), y_train.numpy(), width_vec=[d, width, width, width, len(taus)], lr=lr, epochs=100)
    net_DQRP = DQRP(quantiles=taus)
    net_DQRP.fit(X=x_train.numpy(), y=y_train.numpy(), width_vec=[d+1, width, width, width, 1], activation='ReQU', lr=lr, epochs=100)
    net_cheb = ChebNet(quantiles=taus)
    net_cheb.fit(X=x_train.numpy(), y=y_train.numpy(), phi_layer=[d, width, width, 15], K_layer=[d, width, width, 1], K='q0', lr=lr, epochs=100)

    qr = LinearQR(taus)
    qr.fit(x_train, y_train)
    qrf = QuantileForest(quantiles=taus, nthreads=4)
    qrf.fit(x_train, y_train)
    kqr = KernelQR(quantiles=taus)
    if SIZE <= 512:
        kqr.fit(x_train, y_train)
    splqr = SplineQR(quantiles=taus)
    if d <= 2:
        splqr.fit(x_train, y_train)

    # Test all models
    x_test = torch.rand([10000, d])
    quants = quant_multi(x_test, taus, A, B, model=model, error=error, df=df)

    preds = {}
    preds['QR'] = torch.from_numpy(qr.predict(x_test))
    preds['RF'] = torch.from_numpy(qrf.predict(x_test))
    preds['KQR'] = torch.from_numpy(kqr.predict(x_test)) if SIZE <= 512 else torch.zeros_like(preds['QR'])
    preds['SPL'] = torch.from_numpy(splqr.predict(x_test)) if d <= 2 else torch.zeros_like(preds['QR'])
    preds['cheb'] = net_cheb.predict(x_test, taus)
    preds['DQRP'] = net_DQRP.predict(x_test, taus)
    preds['NQ'] = net_NQ.predict(x_test)
    preds['NC'] = net_NC.predict(x_test)
    preds['DQR'] = net_DQR.predict(x_test)

    # Compute L1 and L2 errors for each method
    for name in method_names:
        L1_dict[name][r, :] = torch.abs(preds[name] - quants).mean(0)
        L2_dict[name][r, :] = torch.pow(preds[name] - quants, 2).mean(0)

    print(f"Replication {r+1} End.")

#%% Organize and save the results

round_dig = 2
tau_values = np.round(taus.squeeze().numpy(), 3)
col_names = ['L1_mean', 'L1_std', 'L2_mean', 'L2_std']

results = {}

for name in method_names:
    L1 = L1_dict[name]
    L2 = L2_dict[name]
    L1_mean = L1.mean(dim=0).detach().numpy().reshape(-1, 1).round(round_dig)
    L1_std = L1.std(dim=0).detach().numpy().reshape(-1, 1).round(round_dig)
    L2_mean = L2.mean(dim=0).detach().numpy().reshape(-1, 1).round(round_dig)
    L2_std = L2.std(dim=0).detach().numpy().reshape(-1, 1).round(round_dig)
    arr = np.concatenate([L1_mean, L1_std, L2_mean, L2_std], axis=1)
    df = pd.DataFrame(arr, index=tau_values, columns=col_names)
    df.index.name = 'tau'
    results[name] = df

# Create output folder if it does not exist
folder = './%s_%s_%d' % (model, error, SIZE)
os.makedirs(folder, exist_ok=True)

# Save each method's result as a CSV file
for name, df in results.items():
    df.to_csv(os.path.join(folder, f'{name}.csv'))

#%% Summarize results for all methods under different sample sizes

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ========== Configurations ==========
model='scenario1';error='scenario1';d=6;
#model='scenario1';error='scenario1';d=6;
#model='scenario1';error='scenario1';d=6;

#model='scenario2';error='t';d=15;
#model='scenario2';error='t';d=15;
#model='scenario2';error='t';d=15;

#model='scenario3';error='sinex_t';d=30;
#model='scenario3';error='sinex_t';d=30;
#model='scenario3';error='sinex_t';d=30;

# You can uncomment other settings as needed
sizes = [2**10, 2**11, 2**12]

methods = ['QR', 'RF', 'KQR', 'SPL', 'DQRP', 'cheb', 'DQR', 'NC', 'NQ']
rename_dict = {
    'QR': 'Linear',
    'cheb': 'Cheb-Net',
    'NQ': 'NQ-Net',
    'NC': 'NC-QR-DQN',
    'SPL': 'Spline',
    'RF': 'Forest',
    'KQR': 'Kernel'
}

taus = np.linspace(0.05, 0.95, 19)
taus_list = [f"{x:.2f}" for x in taus]
save_taus = ["0.05", "0.25", "0.50", "0.75", "0.95"]
tau_labels = [f"$\\tau={tau}$" for tau in save_taus]

# ========== Data Collection Function ==========
def collect_all_stats(sizes, methods, taus_list, rename_dict, folder_template):
    all_l1_mean, all_l1_std, all_l2_mean, all_l2_std = {}, {}, {}, {}
    for size in sizes:
        folder = folder_template.format(model=model, error=error, size=size)
        l1_mean_dict, l1_std_dict, l2_mean_dict, l2_std_dict = {}, {}, {}, {}
        for method in methods:
            file_path = os.path.join(folder, f'{method}.csv')
            if not os.path.exists(file_path):
                print(f'Warning: {file_path} not found, skip.')
                continue
            df = pd.read_csv(file_path, index_col=0)
            col_name = rename_dict.get(method, method)
            l1_mean_dict[col_name] = df.iloc[:, 0].values
            l1_std_dict[col_name]  = df.iloc[:, 1].values
            l2_mean_dict[col_name] = df.iloc[:, 2].values
            l2_std_dict[col_name]  = df.iloc[:, 3].values
        all_l1_mean[size] = pd.DataFrame(l1_mean_dict, index=taus_list)
        all_l1_std[size]  = pd.DataFrame(l1_std_dict,  index=taus_list)
        all_l2_mean[size] = pd.DataFrame(l2_mean_dict, index=taus_list)
        all_l2_std[size]  = pd.DataFrame(l2_std_dict,  index=taus_list)
    return all_l1_mean, all_l1_std, all_l2_mean, all_l2_std

# ========== Table Formatting Functions ==========
def make_metric_table(mean_dict, std_dict, sizes, taus_list):
    """Long table: each row is a tau, each block is a sample size."""
    dfs = []
    for size in sizes:
        mean_df = mean_dict[size].reset_index().rename(columns={'index': '$\\tau$'})
        std_df  = std_dict[size].reset_index().rename(columns={'index': '$\\tau$'})
        mean_df.insert(0, '$N$', int(size))
        std_df.insert(0, '$N$', int(size))
        for method in mean_df.columns[2:]:
            mean_df[method] = mean_df.apply(
                lambda row: f"{row[method]:.2f} ({std_df.loc[row.name, method]:.2f})", axis=1
            )
        dfs.append(mean_df)
    df_all = pd.concat(dfs, axis=0, ignore_index=True)
    return df_all

def make_wide_table(mean_dict, std_dict, sizes, save_taus, tau_labels):
    """Wide table: each row is a method under a sample size, columns are selected taus."""
    rows = []
    for size in sizes:
        mean_df = mean_dict[size].loc[save_taus]
        std_df = std_dict[size].loc[save_taus]
        for method in mean_df.columns:
            row = {'$N$': int(size), 'Method': method}
            for tau, tau_label in zip(save_taus, tau_labels):
                row[tau_label] = f"{mean_df.loc[tau, method]:.2f} ({std_df.loc[tau, method]:.2f})"
            rows.append(row)
    cols = ['$N$', 'Method'] + tau_labels
    df = pd.DataFrame(rows)[cols]
    return df

def merge_cells_by_size(excel_path, sheet_names, block_size):
    """Merge the $N$ column for each block of block_size rows."""
    wb = load_workbook(excel_path)
    for sheet in sheet_names:
        ws = wb[sheet]
        n_rows = ws.max_row
        size_col = 1
        for i in range(2, n_rows+1, block_size):
            ws.merge_cells(start_row=i, start_column=size_col, end_row=min(i+block_size-1, n_rows), end_column=size_col)
    wb.save(excel_path)

# ========== Main Process ==========

folder_template = './{model}_{error}_{size}'

# Collect all statistics
all_l1_mean, all_l1_std, all_l2_mean, all_l2_std = collect_all_stats(
    sizes, methods, taus_list, rename_dict, folder_template
)


summary_dir = './simulation_multivariate'
os.makedirs(summary_dir, exist_ok=True)

# 1. Save long table (all taus)
l1_table = make_metric_table(all_l1_mean, all_l1_std, sizes, taus_list)
l2_table = make_metric_table(all_l2_mean, all_l2_std, sizes, taus_list)
output_excel_long = f'./{model}_{error}_summary_full.xlsx'
with pd.ExcelWriter(output_excel_long) as writer:
    l1_table.to_excel(writer, sheet_name='L1', index=False)
    l2_table.to_excel(writer, sheet_name='L2', index=False)
print(f'Full result saved to Excel file: {output_excel_long}')
merge_cells_by_size(output_excel_long, ['L1', 'L2'], block_size=len(taus_list))

# 2. Save wide table (selected taus)
l1_wide = make_wide_table(all_l1_mean, all_l1_std, sizes, save_taus, tau_labels)
l2_wide = make_wide_table(all_l2_mean, all_l2_std, sizes, save_taus, tau_labels)
output_excel_wide = f'./{model}_{error}_summary_selected.xlsx'
with pd.ExcelWriter(output_excel_wide) as writer:
    l1_wide.to_excel(writer, sheet_name='L1', index=False)
    l2_wide.to_excel(writer, sheet_name='L2', index=False)
print(f'Selected result saved to Excel file: {output_excel_wide}')
merge_cells_by_size(output_excel_wide, ['L1', 'L2'], block_size=len(methods))
